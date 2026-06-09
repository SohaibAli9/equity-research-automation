"""Build the live Excel DCF model with openpyxl.

Forecast and valuation cells are FORMULAS that reference the editable assumption cells,
so the model is genuinely live — change an assumption and Excel recomputes the whole
chain (EBIT→NOPAT→FCF→PV→TV→EV→equity bridge→fair value). Historicals/actuals and the
assumption seed values come from the extraction + dcf truth. `fullCalcOnLoad` forces a
recalculation when the file is opened.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.properties import PageSetupProperties

from pipeline import config
from pipeline.schema import Filing
from . import style as S
from .dcf import compute

# Model sheet year columns
COLS = ["B", "C", "D", "E", "F", "G"]            # 2024E..2029E
YEARS = [2024, 2025, 2026, 2027, 2028, 2029]
FC_COLS = COLS[1:]                                # forecast columns C..G

# Assumptions cell addresses (col B), referenced absolutely in formulas
A = {
    "rf": 2, "erp": 3, "beta": 4, "crp": 5, "ke": 6,
    "kd": 7, "tax": 8, "kd_at": 9, "we": 10, "wd": 11, "wacc": 12, "g": 13,
    "ebit_m": 15, "da_pct": 16, "capex_b": 17, "capex_t": 18, "nwc": 19,
    "rev0": 21, "ni0": 22, "net_margin": 23,
    "g25": 25, "g26": 26, "g27": 27, "g28": 28, "g29": 29,
    "net_debt": 31, "nci": 32, "pref": 33, "shares": 34, "price": 35, "pref_div": 36,
}
def AR(key: str) -> str:                          # absolute Assumptions ref
    return f"Assumptions!$B${A[key]}"


def _assumptions_sheet(wb, R):
    ws = wb.create_sheet("Assumptions")
    S.widths(ws, {"A": 34, "B": 16})
    S.put(ws, "A1", "ASSUMPTIONS  —  edit blue cells; the model recomputes", font=S.F_TITLE)
    drv, val = R["drivers"], R["valuation"]

    # input rows (blue): (row, label, value, fmt)
    inputs = [
        (A["rf"], "Risk-free rate (PH 10Y)", R["_a"]["risk_free_rate"], S.PCT),
        (A["erp"], "Mature-market ERP", R["_a"]["mature_market_erp"], S.PCT),
        (A["beta"], "Beta (observed)", R["_a"]["beta"], "0.00"),
        (A["crp"], "Country risk premium", R["_a"]["country_risk_premium"], S.PCT),
        (A["kd"], "Pre-tax cost of debt", R["_a"]["pretax_cost_of_debt"], S.PCT),
        (A["tax"], "Tax rate", drv["tax_rate"], S.PCT),
        (A["we"], "Equity weight (target)", R["_a"]["target_equity_weight"], S.PCT),
        (A["wd"], "Debt weight (target)", R["_a"]["target_debt_weight"], S.PCT),
        (A["g"], "Terminal growth", drv["terminal_growth"], S.PCT),
        (A["ebit_m"], "EBIT margin", drv["ebit_margin"], S.PCT),
        (A["da_pct"], "D&A % of revenue", drv["da_pct"], S.PCT),
        (A["capex_b"], "CapEx % revenue (base)", drv["capex_pct_base"], S.PCT),
        (A["capex_t"], "CapEx % revenue (terminal)", drv["capex_pct_terminal"], S.PCT),
        (A["nwc"], "Δ NWC % of Δrevenue", R["_a"]["forecast_drivers"]["nwc_change_pct_revenue"], S.PCT),
        (A["rev0"], "Base revenue FY2024E (₱000)", R["projections"]["revenue"][2024], S.NUM),
        (A["ni0"], "Base parent net income FY2024E (₱000)", R["_ni0"], S.NUM),
        (A["g25"], "Revenue growth 2025", R["_a"]["forecast_drivers"]["revenue_growth"][2025], S.PCT),
        (A["g26"], "Revenue growth 2026", R["_a"]["forecast_drivers"]["revenue_growth"][2026], S.PCT),
        (A["g27"], "Revenue growth 2027", R["_a"]["forecast_drivers"]["revenue_growth"][2027], S.PCT),
        (A["g28"], "Revenue growth 2028", R["_a"]["forecast_drivers"]["revenue_growth"][2028], S.PCT),
        (A["g29"], "Revenue growth 2029", R["_a"]["forecast_drivers"]["revenue_growth"][2029], S.PCT),
        (A["net_debt"], "Net debt (₱000)", val["net_debt"], S.NUM),
        (A["nci"], "Non-controlling interests (₱000)", val["nci"], S.NUM),
        (A["pref"], "Preferred (book, ₱000)", val["preferred"], S.NUM),
        (A["shares"], "Common shares outstanding", val["common_shares"], "#,##0"),
        (A["price"], "Current price (₱)", val["current_price"], S.PRICE),
        (A["pref_div"], "Preferred dividend (₱000/yr)", R["_a"]["forecast_drivers"]["preferred_dividend_annual"], S.NUM),
    ]
    for row, label, value, fmt in inputs:
        S.put(ws, f"A{row}", label)
        S.put(ws, f"B{row}", value, fmt=fmt, font=S.F_INPUT)

    # formula cells (black)
    formulas = [
        (A["ke"], "Cost of equity (CAPM+CRP)",
         f"=$B${A['rf']}+$B${A['beta']}*$B${A['erp']}+$B${A['crp']}", S.PCT),
        (A["kd_at"], "After-tax cost of debt",
         f"=$B${A['kd']}*(1-$B${A['tax']})", S.PCT),
        (A["wacc"], "WACC",
         f"=$B${A['we']}*$B${A['ke']}+$B${A['wd']}*$B${A['kd_at']}", S.PCT),
        (A["net_margin"], "Net margin (parent)",
         f"=$B${A['ni0']}/$B${A['rev0']}", S.PCT),
    ]
    for row, label, formula, fmt in formulas:
        S.put(ws, f"A{row}", label, bold=True)
        S.put(ws, f"B{row}", formula, fmt=fmt, font=S.F_BOLD)
    return ws


def _model_sheet(wb):
    ws = wb.create_sheet("Model")
    S.widths(ws, {"A": 26, **{c: 13 for c in COLS}})
    S.put(ws, "A1", "THREE-STATEMENT MODEL  (₱ thousands)", font=S.F_TITLE)
    # header
    for col, yr in zip(COLS, YEARS):
        tag = "E" if yr <= 2027 else "F"
        if yr == 2024:
            tag = "E"
        S.put(ws, f"{col}2", f"{yr}{tag}", font=S.F_HDR, fill=S.FILL_HDR, align="center")
    S.put(ws, "A2", "₱000 / FY", font=S.F_HDR, fill=S.FILL_HDR)

    def row(r, label, builder, fmt=S.NUM, bold=False):
        S.put(ws, f"A{r}", label, bold=bold)
        for i, col in enumerate(COLS):
            S.put(ws, f"{col}{r}", builder(i, col), fmt=fmt, bold=bold)

    # Revenue: base = assumption; forecast = prior * (1+growth)
    gkeys = ["g25", "g26", "g27", "g28", "g29"]
    def revenue(i, col):
        if i == 0:
            return f"={AR('rev0')}"
        return f"={COLS[i-1]}4*(1+{AR(gkeys[i-1])})"
    row(4, "Revenue", revenue, bold=True)
    row(5, "EBIT", lambda i, c: f"={c}4*{AR('ebit_m')}")
    row(6, "NOPAT", lambda i, c: f"={c}5*(1-{AR('tax')})")
    row(7, "D&A", lambda i, c: f"={c}4*{AR('da_pct')}")

    # CapEx % fades linearly from base (2024) to terminal (2029)
    def capex_pct(i, col):
        if i == 0:
            return f"={AR('capex_b')}"
        frac = i / (len(COLS) - 1)
        return f"={AR('capex_b')}+({AR('capex_t')}-{AR('capex_b')})*{frac}"
    row(8, "CapEx % revenue", capex_pct, fmt=S.PCT)
    row(9, "CapEx", lambda i, c: f"={c}4*{c}8")
    row(10, "Δ NWC", lambda i, c: "=0" if i == 0 else f"=({c}4-{COLS[i-1]}4)*{AR('nwc')}")
    row(11, "Free cash flow (FCFF)", lambda i, c: f"={c}6+{c}7-{c}9-{c}10", bold=True)
    ws[f"A11"].fill = S.FILL_TOTAL

    # Net income (parent) + EPS for the note's financial summary
    row(13, "Net income (parent)",
        lambda i, c: f"={AR('ni0')}" if i == 0 else f"={c}4*{AR('net_margin')}")
    row(14, "EPS (₱)", lambda i, c: f"=({c}13-{AR('pref_div')})*1000/{AR('shares')}", fmt=S.NUM2)
    row(15, "EPS growth",
        lambda i, c: "" if i == 0 else f"={c}14/{COLS[i-1]}14-1", fmt=S.PCT)
    return ws


def _valuation_sheet(wb, R):
    ws = wb.create_sheet("Valuation")
    S.widths(ws, {"A": 28, "B": 15, **{c: 13 for c in FC_COLS}})
    S.put(ws, "A1", "DCF VALUATION", font=S.F_TITLE)
    S.put(ws, "A2", "WACC"); S.put(ws, "B2", f"={AR('wacc')}", fmt=S.PCT, bold=True)
    S.put(ws, "A3", "Terminal growth"); S.put(ws, "B3", f"={AR('g')}", fmt=S.PCT)

    # discount table over forecast years (cols C..G = 2025..2029)
    S.put(ws, "A5", "FY", font=S.F_HDR, fill=S.FILL_HDR)
    for col, yr in zip(FC_COLS, YEARS[1:]):
        tag = "E" if yr <= 2027 else "F"
        S.put(ws, f"{col}5", f"{yr}{tag}", font=S.F_HDR, fill=S.FILL_HDR, align="center")
    S.put(ws, "A6", "Free cash flow")
    S.put(ws, "A7", "Discount factor")
    S.put(ws, "A8", "PV of FCF")
    for t, col in enumerate(FC_COLS, 1):
        S.put(ws, f"{col}6", f"=Model!{col}11", fmt=S.NUM)
        S.put(ws, f"{col}7", f"=1/(1+$B$2)^{t}", fmt="0.000")
        S.put(ws, f"{col}8", f"={col}6*{col}7", fmt=S.NUM)

    chain = [
        (10, "Σ PV of explicit FCF", "=SUM(C8:G8)", S.NUM, True),
        (11, "Terminal-year FCF", "=G6", S.NUM, False),
        (12, "Terminal value (Gordon)", "=B11*(1+B3)/(B2-B3)", S.NUM, False),
        (13, "PV of terminal value", f"=B12/(1+B2)^{len(FC_COLS)}", S.NUM, False),
        (14, "Enterprise value", "=B10+B13", S.NUM, True),
        (15, "Less: net debt", f"=-{AR('net_debt')}", S.NUM, False),
        (16, "Less: NCI", f"=-{AR('nci')}", S.NUM, False),
        (17, "Less: preferred", f"=-{AR('pref')}", S.NUM, False),
        (18, "Common equity value", "=B14+B15+B16+B17", S.NUM, True),
        (19, "Common shares", f"={AR('shares')}", "#,##0", False),
        (20, "FAIR VALUE / SHARE (₱)", "=B18*1000/B19", S.PRICE, True),
        (21, "Current price (₱)", f"={AR('price')}", S.PRICE, False),
        (22, "Upside / (downside)", "=B20/B21-1", S.PCT, True),
        (23, "Recommendation",
         '=IF(B22>0.1,"BUY",IF(B22>=0,"HOLD","UNDERPERFORM"))', None, True),
    ]
    for r, label, formula, fmt, bold in chain:
        S.put(ws, f"A{r}", label, bold=bold)
        c = S.put(ws, f"B{r}", formula, fmt=fmt, bold=bold)
        if r in (20, 23):
            c.fill = S.FILL_RESULT
            ws[f"A{r}"].fill = S.FILL_RESULT

    # Sensitivity grid (snapshot values from the Python truth)
    sens = R["sensitivity"]
    S.put(ws, "A26", "SENSITIVITY — fair value/share (₱)", bold=True)
    S.put(ws, "A27", "WACC ↓  /  g →", font=S.F_HDR, fill=S.FILL_HDR)
    for j, g in enumerate(sens["growth_axis"]):
        S.put(ws, f"{FC_COLS[j]}27", g, fmt=S.PCT, font=S.F_HDR, fill=S.FILL_HDR, align="center")
    for i, wv in enumerate(sens["wacc_axis"]):
        S.put(ws, f"A{28+i}", wv, fmt=S.PCT)
        for j, _ in enumerate(sens["growth_axis"]):
            S.put(ws, f"{FC_COLS[j]}{28+i}", sens["grid"][i][j], fmt=S.NUM2)
    return ws


def _summary_sheet(wb):
    ws = wb.create_sheet("Financial summary", 0)
    S.widths(ws, {"A": 26, "B": 14, "C": 14, "D": 14, "E": 14})
    S.put(ws, "A1", "FINANCIAL SUMMARY", font=S.F_TITLE)
    S.put(ws, "A2", "Rating / target — see Valuation tab", font=S.F_BOLD)
    S.put(ws, "A3", "Recommendation"); S.put(ws, "B3", "=Valuation!B23", bold=True, fill=S.FILL_RESULT)
    S.put(ws, "A4", "Fair value / share (₱)"); S.put(ws, "B4", "=Valuation!B20", fmt=S.PRICE, bold=True)
    S.put(ws, "A5", "Upside"); S.put(ws, "B5", "=Valuation!B22", fmt=S.PCT, bold=True)

    cols = ["B", "C", "D", "E"]
    S.put(ws, "A7", "₱000 / FY", font=S.F_HDR, fill=S.FILL_HDR)
    for col, yr in zip(cols, [2024, 2025, 2026, 2027]):
        S.put(ws, f"{col}7", f"{yr}{'E' if yr <= 2027 else 'F'}", font=S.F_HDR, fill=S.FILL_HDR, align="center")
    line = [(8, "Revenue", 4, S.NUM), (9, "EBIT", 5, S.NUM), (10, "NOPAT", 6, S.NUM),
            (11, "Free cash flow", 11, S.NUM), (12, "Net income (parent)", 13, S.NUM),
            (13, "EPS (₱)", 14, S.NUM2)]
    for r, label, mrow, fmt in line:
        S.put(ws, f"A{r}", label)
        for col, mcol in zip(cols, ["B", "C", "D", "E"]):
            S.put(ws, f"{col}{r}", f"=Model!{mcol}{mrow}", fmt=fmt)
    return ws


def _interim_sheet(wb, F: Filing):
    ws = wb.create_sheet("Interim")
    S.widths(ws, {"A": 44, "B": 13, "C": 13, "D": 13, "E": 13})
    S.put(ws, "A1", "INTERIM RESULTS — income statement (₱000, from filing)", font=S.F_TITLE)
    hdr = ["Q3 2024", "Q3 2023", "9M 2024", "9M 2023"]
    for col, h in zip(["B", "C", "D", "E"], hdr):
        S.put(ws, f"{col}2", h, font=S.F_HDR, fill=S.FILL_HDR, align="center")
    pk = config.INCOME_PERIODS
    for i, li in enumerate(F.income.line_items, 3):
        S.put(ws, f"A{i}", li.label)
        for col, p in zip(["B", "C", "D", "E"], pk):
            v = li.values.get(p)
            if v is not None:
                S.put(ws, f"{col}{i}", v, fmt=S.NUM2 if "eps" in (li.account or "") else S.NUM)
    return ws


def _segments_sheet(wb, F: Filing):
    ws = wb.create_sheet("Segments")
    S.widths(ws, {"A": 40, "B": 14, "C": 14, "D": 13, "E": 14, "F": 14})
    S.put(ws, "A1", "SEGMENTS — Note 15, 9M 2024 (₱000, from filing)", font=S.F_TITLE)
    if not F.segments:
        return ws
    hdr = ["Manila Conc.", "Domestic", "Foreign", "Consol. adj.", "Consolidated"]
    for col, h in zip(["B", "C", "D", "E", "F"], hdr):
        S.put(ws, f"{col}2", h, font=S.F_HDR, fill=S.FILL_HDR, align="center")
    for i, rowd in enumerate(F.segments.rows, 3):
        S.put(ws, f"A{i}", rowd.label)
        for col, seg in zip(["B", "C", "D", "E", "F"], F.segments.segments):
            v = rowd.values.get(seg)
            if v is not None:
                S.put(ws, f"{col}{i}", v, fmt=S.NUM)
    return ws


def _charts_sheet(wb):
    from openpyxl.chart import BarChart, LineChart, Reference, Series
    ws = wb.create_sheet("Charts")
    S.put(ws, "A1", "CHARTS", font=S.F_TITLE)
    model = wb["Model"]
    cats = Reference(model, min_col=2, max_col=7, min_row=2, max_row=2)

    lc = LineChart()
    lc.title = "Revenue & Free Cash Flow (₱000)"
    lc.height, lc.width = 8, 16
    for r in (4, 11):
        s = Series(Reference(model, min_col=2, max_col=7, min_row=r, max_row=r),
                   title=model[f"A{r}"].value)
        lc.series.append(s)
    lc.set_categories(cats)
    ws.add_chart(lc, "A3")

    bc = BarChart()
    bc.title = "EPS (₱)"
    bc.height, bc.width = 8, 16
    bc.add_data(Reference(model, min_col=2, max_col=7, min_row=14, max_row=14), titles_from_data=False)
    bc.set_categories(cats)
    bc.legend = None
    ws.add_chart(bc, "A20")
    return ws


def build() -> Workbook:
    R = compute()
    # enrich R with a couple of values workbook needs
    import yaml
    R["_a"] = yaml.safe_load((config.CONFIG_DIR / "assumptions.yaml").read_text())["valuation"]
    R["_a"]["forecast_drivers"] = yaml.safe_load(
        (config.CONFIG_DIR / "assumptions.yaml").read_text())["forecast_drivers"]
    F = Filing.model_validate_json((config.output_dir_for() / "financials.json").read_text())
    R["_ni0"] = R["projections"]["net_income"][R["years"][0]]   # corrected FY2024E base NI

    wb = Workbook()
    wb.remove(wb.active)
    _summary_sheet(wb)
    _assumptions_sheet(wb, R)
    _model_sheet(wb)
    _valuation_sheet(wb, R)
    _interim_sheet(wb, F)
    _segments_sheet(wb, F)
    _charts_sheet(wb)
    wb.calculation.fullCalcOnLoad = True

    # Print layout: each tab fits one page wide (clean PDF export / printing)
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B3"
    return wb
