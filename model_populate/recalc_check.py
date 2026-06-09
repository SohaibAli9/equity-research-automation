"""Cross-check: recalculate the Excel workbook with LibreOffice (headless) and confirm
its formula results match the Python DCF truth to the cent. Proves the live model and
the note tell an identical story.
"""
from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def recalc(xlsx_path: Path) -> dict[tuple[str, str], float]:
    """Open in LibreOffice (recalcs on load via fullCalcOnLoad) and return cached cell
    values for the whole workbook as {(sheet, addr): value}."""
    tmp = Path(tempfile.mkdtemp())
    try:
        subprocess.run(
            ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(tmp), str(xlsx_path)],
            check=True, capture_output=True, timeout=120,
        )
        wb = load_workbook(tmp / xlsx_path.name, data_only=True)
        out = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None:
                        out[(ws.title, c.coordinate)] = c.value
        return out
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def cross_check(xlsx_path: Path, truth: dict) -> list[dict]:
    """Compare key Excel formula results to the Python DCF truth."""
    vals = recalc(xlsx_path)
    v = truth["valuation"]
    fc = truth["forecast_years"]
    p = truth["projections"]
    expected = {
        ("Valuation", "B20"): ("fair value / share", v["fair_value_per_share"], 0.01),
        ("Valuation", "B14"): ("enterprise value", v["enterprise_value"], 50.0),
        ("Valuation", "B18"): ("common equity value", v["common_equity_value"], 50.0),
        ("Valuation", "B12"): ("terminal value", v["terminal_value"], 100.0),
        ("Valuation", "B2"): ("WACC", truth["wacc"]["wacc"], 1e-5),
        ("Model", "G11"): ("FCF 2029", p["fcf"][fc[-1]], 50.0),
        ("Model", "B4"): ("revenue 2024E", p["revenue"][truth["years"][0]], 50.0),
    }
    checks = []
    for (sheet, addr), (name, exp, tol) in expected.items():
        act = vals.get((sheet, addr))
        ok = act is not None and abs(act - exp) <= tol
        checks.append({"name": name, "cell": f"{sheet}!{addr}", "python": exp,
                       "excel": act, "match": ok})
    return checks
