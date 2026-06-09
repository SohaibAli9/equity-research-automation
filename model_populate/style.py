"""Number formats, fonts, and cell helpers for the Excel model.

Sell-side convention: blue font = hardcoded input/assumption cells; black = formulas.
Negatives in parentheses. Forecast columns flagged E/F in headers.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Number formats
NUM = "#,##0;(#,##0)"            # ₱ thousands
NUM2 = "#,##0.00;(#,##0.00)"
PRICE = '"₱"#,##0.00'
PCT = "0.0%"
PCT2 = "0.00%"
MULT = '0.0"x"'

# Fonts
F_INPUT = Font(color="0000CC")                       # blue = input
F_FORMULA = Font(color="000000")                     # black = formula
F_BOLD = Font(bold=True)
F_TITLE = Font(bold=True, size=13)
F_SECTION = Font(bold=True, color="FFFFFF")
F_HDR = Font(bold=True)

# Fills
FILL_SECTION = PatternFill("solid", fgColor="1F3864")   # dark navy
FILL_HDR = PatternFill("solid", fgColor="D9E1F2")       # light blue
FILL_TOTAL = PatternFill("solid", fgColor="F2F2F2")
FILL_RESULT = PatternFill("solid", fgColor="FFF2CC")    # highlight result

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TOP = Border(top=Side(style="thin", color="000000"))


def put(ws, addr, value, *, fmt=None, font=None, fill=None, bold=False,
        align=None, border=False):
    c = ws[addr]
    c.value = value
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = Font(bold=True, color=(font.color.rgb if font and font.color else "000000"))
    elif font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = Alignment(horizontal=align)
    if border:
        c.border = BORDER
    return c


def section(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION


def widths(ws, mapping: dict[str, int]):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w
